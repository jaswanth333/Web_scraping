from bs4 import BeautifulSoup
import  requests
import  openpyxl

excel=openpyxl.Workbook()
print(excel.sheetnames)
sheet=excel.active
sheet.title="IMDB list"
print(excel.sheetnames)

sheet.append(["Movie_Rank","Movie_Name","Year","MovieRating"])


try:
    url=requests.get("https://www.imdb.com/chart/top/")
    url.raise_for_status()
    
except Exception as e:
    print(e)

doc=BeautifulSoup(url.text,"html.parser")
movies=doc.find("tbody",class_="lister-list").find_all("tr")

for movie in movies:
    rank=int(movie.find("td",class_="titleColumn").get_text(strip=True).split(".")[0])
    name=movie.find("td",class_="titleColumn").a.text
    year=int(movie.find("td",class_="titleColumn").span.text.strip("()"))
    rating=movie.find("td",class_="ratingColumn imdbRating").strong.text
    

    print(rank,name,year,rating)
    sheet.append([rank,name,year,rating])

excel.save("D:\Jupyter\Python\Web Scraping\IMDB_list.xlsx")